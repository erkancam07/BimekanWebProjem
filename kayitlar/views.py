from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.urls import reverse_lazy
from django.views.generic import UpdateView, DeleteView,CreateView
from django.utils import timezone # Django'nun timezone'unu kullanıyoruz
from django.db.models import Q, Sum 
from datetime import datetime, date, time, timedelta # datetime sınıfı zaten burada import edilmiş
from calendar import monthrange
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin

from .models import Misafir, Islem, Yatak, IslemTuru, YoklamaDurumu, YoklamaKaydi
from .forms import MisafirKayitForm, MisafirIslemForm, MisafirGuncelleForm, GunlukYoklamaForm

from django.contrib.auth.forms import UserCreationForm
from django.views import generic

from django.http import HttpResponse, JsonResponse
from django.template.loader import render_to_string
<<<<<<< HEAD
import math
=======
>>>>>>> 3df35d64b63d79cb98d0843a2f23eefade12dd17

# Excel dışa aktarımı için gerekli importlar
from django.http import HttpResponse
import openpyxl 
# import pytz # BU SATIRI ARTIK GEREKLİ DEĞİL, SİLİN VEYA YORUM SATIRI YAPIN!
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from .models import SosyalGuvence
from .forms import SosyalGuvenceForm
from .models import Kurum

<<<<<<< HEAD
from .forms import StokEkleForm,GiyimIslemForm
from .models import GiyimUrunu

from django.db import IntegrityError # unique_together hatasını yakalamak için
from django.forms import ValidationError # Formun clean metodundan gelen hata için

# --- 1. Yeni Ürün Tanımlama View'ı (Mevcut 'stok_ekle' view'ının yerine geçecek ve adı değişecek) ---
# Bu view, sadece yeni bir GiyimUrunu (örn: Terlik-Ayakkabı kombinasyonu) tanımlamak içindir.
def yeni_urun_tanimla(request):
    form = StokEkleForm(request.POST or None)
    if request.method == "POST":
        if form.is_valid():
            try:
                # Burada sadece GiyimUrunu objesini kaydediyoruz.
                # 'mevcut_adet' alanı, ürün ilk tanımlandığında başlangıç stoğu olabilir.
                form.save()
                # Başarılı kayıttan sonra kullanıcıyı bilgilendirip, belki stok raporuna yönlendirebiliriz.
                # messages.success(request, "Yeni ürün başarıyla tanımlandı!") # Django mesaj sistemi kullanıyorsanız
                return redirect("stok_raporu") # Lütfen bu URL'in tanımlı olduğundan emin olun
            except IntegrityError: # unique_together kısıtlamasından kaynaklanan hatayı yakala
                # Eğer formun clean metodu bu hatayı yakalamazsa, burada yakalarız
                form.add_error(None, "Bu ürün adı ve kategori kombinasyonu zaten mevcut. Lütfen mevcut ürünü güncelleyin veya 'Stok Hareketi Ekle' formunu kullanın.")
            except Exception as e: # Diğer olası hataları yakala
                form.add_error(None, f"Beklenmeyen bir hata oluştu: {e}")
        # else: form.is_valid() değilse, hatalar form objesinde saklanır ve şablona gönderilir
    return render(request, "bimekan/stok_ekle.html", {"form": form})


# --- 2. Stok Hareketi Ekleme View'ı (YEPYENİ BİR FONKSİYON) ---
# Bu view, mevcut GiyimUrunu'nun stok miktarını giriş/çıkış yaparak güncellemek içindir.
def stok_hareket_ekle(request):
    form = GiyimIslemForm(request.POST or None)
    if request.method == "POST":
        if form.is_valid():
            try:
                # GiyimIslem kaydedildiğinde, GiyimIslem modelindeki save() metodu (models.py'de)
                # ilgili GiyimUrunu'nun 'mevcut_adet'ini otomatik olarak güncelleyecektir.
                form.save()
                # messages.success(request, "Stok hareketi başarıyla kaydedildi!") # Django mesaj sistemi kullanıyorsanız
                return redirect("stok_raporu") # İşlem sonrası rapor sayfasına yönlendir
            except ValidationError as e: # Formun clean metodundan (GiyimIslemForm'dan) gelen doğrulama hatalarını yakala
                form.add_error(None, e) # Hata mesajını forma ekle
            except Exception as e: # Diğer olası hataları yakala
                form.add_error(None, f"Bir hata oluştu: {e}")
        # else: form.is_valid() değilse, hatalar form objesinde saklanır ve şablona gönderilir
    return render(request, "bimekan/stok_hareket_ekle.html", {"form": form}) # Bu şablonu oluşturmanız gerekecek

def stok_raporu(request):
    genel_stok = GiyimUrunu.objects.all().order_by('ad__isim', 'kategori')

    stok_girisleri = GiyimIslem.objects.filter(islem_turu='Giriş').order_by('-tarih')
    
    # Kuruma iade işlemleri (GiyimIslem modelinden)
    stok_kurum_iadeleri = GiyimIslem.objects.filter(islem_turu='Çıkış').order_by('-tarih') # Yeni eklendi

    ayni_yardim_turu = None
    try:
        ayni_yardim_turu = IslemTuru.objects.get(ad__iexact='Ayni Yardım')
    except IslemTuru.DoesNotExist:
        pass

    stok_misafir_cikis_islemleri = Islem.objects.none() # İsim daha açıklayıcı oldu
    if ayni_yardim_turu:
        stok_misafir_cikis_islemleri = Islem.objects.filter(
            islem_turu=ayni_yardim_turu,
            urun__isnull=False,
            miktar__isnull=False
        ).order_by('-islem_zamani')
    
    tum_stok_hareketleri = []

    for giris in stok_girisleri:
        tum_stok_hareketleri.append({
            'tarih': giris.tarih,
            'urun_ad': giris.urun.ad.isim,
            'urun_kategori': giris.urun.kategori,
            'miktar': giris.miktar,
            'islem_turu': 'Stok Girişi', # Daha açıklayıcı
            'ilgili_kisi': giris.kaynak_firma if giris.kaynak_firma else 'Belirtilmemiş',
            'aciklama': giris.aciklama,
            'model_type': 'GiyimIslem'
        })
    
    # Kuruma iadeleri ekle
    for iade in stok_kurum_iadeleri:
        tum_stok_hareketleri.append({
            'tarih': iade.tarih,
            'urun_ad': iade.urun.ad.isim,
            'urun_kategori': iade.urun.kategori,
            'miktar': iade.miktar,
            'islem_turu': 'Kuruma İade', # Daha açıklayıcı
            'ilgili_kisi': iade.kurum.kurum_adi if iade.kurum else 'Belirtilmemiş', # Kurum adını çek
            'aciklama': iade.aciklama,
            'model_type': 'GiyimIslem'
        })

    for cikis in stok_misafir_cikis_islemleri:
        tum_stok_hareketleri.append({
            'tarih': cikis.islem_zamani,
            'urun_ad': cikis.urun.ad.isim,
            'urun_kategori': cikis.urun.kategori,
            'miktar': cikis.miktar,
            'islem_turu': 'Ayni Yardım', # Daha açıklayıcı
            'ilgili_kisi': f"{cikis.misafir.ad} {cikis.misafir.soyad}",
            'aciklama': cikis.aciklama,
            'model_type': 'Islem'
        })

    tum_stok_hareketleri.sort(key=lambda x: x['tarih'], reverse=True)

    context = {
        'genel_stok': genel_stok,
        'son_islemler': tum_stok_hareketleri[:50],
    }
    return render(request, "bimekan/stok_raporu.html", context)

""" def stok_raporu(request):
    # Genel stok durumu (her bir GiyimUrunu kaydı, kendi toplam mevcut adetini tutar)
    genel_stok = GiyimUrunu.objects.all().order_by('ad__isim', 'kategori')

    # 1. GiyimIslem'den gelen stok girişleri
    # GiyimIslem'deki islem_turu alanı büyük ihtimalle 'Giriş' veya benzeri bir string tutuyordur.
    stok_girisleri = GiyimIslem.objects.filter(islem_turu='Giriş').order_by('-tarih')
    
    # 2. Islem modelinden gelen ayni yardım çıkışları
    ayni_yardim_turu = None
    try:
        ayni_yardim_turu = IslemTuru.objects.get(ad__iexact='Ayni Yardım')
    except IslemTuru.DoesNotExist:
        pass

    stok_cikis_islemleri = Islem.objects.none()
    if ayni_yardim_turu:
        stok_cikis_islemleri = Islem.objects.filter(
            islem_turu=ayni_yardim_turu,
            urun__isnull=False,
            miktar__isnull=False
        ).order_by('-islem_zamani')
    
    tum_stok_hareketleri = []

    for giris in stok_girisleri:
        tum_stok_hareketleri.append({
            'tarih': giris.tarih,
            'urun_ad': giris.urun.ad.isim,
            'urun_kategori': giris.urun.kategori,
            'miktar': giris.miktar,
            'islem_turu': 'Giriş',
            'ilgili_kisi': giris.kaynak_firma if giris.kaynak_firma else 'N/A', # <<< BURASI GÜNCELLENDİ
            'aciklama': giris.aciklama,
            'model_type': 'GiyimIslem'
        })

    for cikis in stok_cikis_islemleri:
        tum_stok_hareketleri.append({
            'tarih': cikis.islem_zamani,
            'urun_ad': cikis.urun.ad.isim,
            'urun_kategori': cikis.urun.kategori,
            'miktar': cikis.miktar,
            'islem_turu': 'Ayni Yardım Çıkışı',
            'ilgili_kisi': f"{cikis.misafir.ad} {cikis.misafir.soyad}",
            'aciklama': cikis.aciklama,
            'model_type': 'Islem'
        })

    tum_stok_hareketleri.sort(key=lambda x: x['tarih'], reverse=True)

    context = {
        'genel_stok': genel_stok,
        'son_islemler': tum_stok_hareketleri[:50],
    }
    return render(request, "bimekan/stok_raporu.html", context) """

=======
from .forms import StokEkleForm
from .models import GiyimUrunu

def stok_raporu(request):
    urunler = GiyimUrunu.objects.all().order_by("ad")
    return render(request, "bimekan/stok_raporu.html", {"urunler": urunler})


def stok_ekle(request):
    form = StokEkleForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        form.save()
        return redirect("stok_raporu")  # Girişten sonra rapora yönlendirme
    return render(request, "bimekan/stok_ekle.html", {"form": form})
>>>>>>> 3df35d64b63d79cb98d0843a2f23eefade12dd17

@login_required
def yardim_gecmisi(request, urun_id):
    urun = get_object_or_404(GiyimUrunu, pk=urun_id)
<<<<<<< HEAD

    try:
        ayni_yardim_turu = IslemTuru.objects.get(ad__iexact='Ayni Yardım')
    except IslemTuru.DoesNotExist:
        ayni_yardim_turu = None
        yardimlar = Islem.objects.none()
        toplam_miktar = 0
        yardim_sayisi = 0 # Yeni eklendi
        kisi_sayisi = 0 # Yeni eklendi
    
    if ayni_yardim_turu:
        yardimlar = Islem.objects.filter(
            islem_turu=ayni_yardim_turu,
            urun=urun,
            miktar__isnull=False,
            miktar__gt=0
        ).order_by("-islem_zamani")

        toplam_miktar = yardimlar.aggregate(toplam=Sum("miktar"))["toplam"] or 0
        
        # İşlem sayısını hesapla
        yardim_sayisi = yardimlar.count() # Yeni eklendi

        # Farklı kişi sayısını hesapla
        kisi_sayisi = yardimlar.values('misafir').distinct().count() # Yeni eklendi
    
=======
    yardimlar = GiyimIslem.objects.filter(urun=urun, islem_turu="Çıkış").order_by("-id")
    toplam_miktar = yardimlar.aggregate(toplam=Sum("miktar"))["toplam"] or 0
>>>>>>> 3df35d64b63d79cb98d0843a2f23eefade12dd17
    context = {
        "urun": urun,
        "yardimlar": yardimlar,
        "toplam_miktar": toplam_miktar,
<<<<<<< HEAD
        "yardim_sayisi": yardim_sayisi, # Yeni eklendi
        "kisi_sayisi": kisi_sayisi,     # Yeni eklendi
=======
>>>>>>> 3df35d64b63d79cb98d0843a2f23eefade12dd17
        "title": f"{urun.ad} Yardım Geçmişi",
    }
    return render(request, "bimekan/yardim_gecmisi.html", context)

# Başvuru formu çıktı view'ı
def basvuru_formu_cikti_view(request, pk):
    misafir = None # Başlangıçta misafir objesini None olarak ayarla

    if pk != 0: # Eğer URL'den gelen 'pk' 0 değilse (yani bir misafir ID'si varsa)
        try:
            misafir = Misafir.objects.get(pk=pk) # Belirtilen pk'ye sahip misafiri çek
        except Misafir.DoesNotExist:
            # Misafir bulunamazsa, yine de 404 yerine boş bir form gösteririz.
            # İsteğe bağlı olarak burada bir loglama yapabilir veya mesaj gösterebiliriz.
            pass # misafir objesi None olarak kalacak

    # current_section'ı göndererek sidebar'da ilgili menü öğesinin aktif olmasını sağla
    context = {
        'misafir': misafir,
        'current_section': 'basvuru-formu' 
    }
    return render(request, 'bimekan/basvuru_formu.html', context)

# Yeni Excel Export View'ı
def misafir_export_excel(request, liste_turu):
    # Django'nun zaman dilimi duyarlı mevcut zamanı
    # Not: Kalış süresi kaldırıldığı için bu 'current_time' değişkeni artık zorunlu değil
    # ancak diğer fonksiyonlarda kullanılıyor olabilir, şimdilik tutalım.
    current_time = timezone.now() 

    misafirler = Misafir.objects.all().order_by('-giris_tarihi')
    if liste_turu == 'aktifler':
        misafirler = misafirler.filter(durum='AKTIF')
    elif liste_turu == 'pasifler':
        misafirler = misafirler.filter(durum='PASIF')
    # 'tumu' durumu için filtreleme yok

    query = request.GET.get('q')
    if query:
        misafirler = misafirler.filter(
            Q(ad__icontains=query) |
            Q(soyad__icontains=query) |
            Q(tc_kimlik_no__icontains=query) |
            Q(dosya_no__icontains=query)
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Dosya başlığı daha genel olacak
    if liste_turu == 'aktifler':
        ws.title = "Kalan Misafirler"
    elif liste_turu == 'pasifler':
        ws.title = "Çıkış Yapan Misafirler"
    else:
        ws.title = "Tüm Misafirler"


    # Excel Başlıkları (İstenilen değişikliklere göre ayarlandı)
    headers = [
        "Dosya No", "Adı", "Soyadı", "TC Kimlik No", "Telefon", "Giriş Tarihi",
        "Durum", "Doğum Tarihi", "Doğum Yeri", "Yatak No", "Adres", "Sosyal Güvence"
    ]
    ws.append(headers)

    # Başlık stilini ayarla
    header_font = Font(bold=True, color="000000") # Yazı rengi hala beyaz kalabilir, okunurluk için iyi
    header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid") # <-- BU SATIR DEĞİŞTİRİLDİ
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )

    for col_num, header_text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_num)].width = 15 # Varsayılan genişlik

    # Verileri ekle (İstenilen değişikliklere göre ayarlandı)
    row_num = 2
    for misafir in misafirler:
        ws.cell(row=row_num, column=1, value=misafir.dosya_no)
        ws.cell(row=row_num, column=2, value=misafir.ad) # Adı ayrı sütunda
        ws.cell(row=row_num, column=3, value=misafir.soyad) # Soyadı ayrı sütunda
        ws.cell(row=row_num, column=4, value=misafir.tc_kimlik_no)
        ws.cell(row=row_num, column=5, value=misafir.telefon)
        ws.cell(row=row_num, column=6, value=misafir.giris_tarihi.strftime('%d.%m.%Y %H:%M') if misafir.giris_tarihi else '-')
        # ws.cell(row=row_num, column=7, value=misafir.cikis_tarihi.strftime('%d.%m.%Y %H:%M') if misafir.cikis_tarihi else '-') # Çıkış tarihi kaldırıldı
        ws.cell(row=row_num, column=7, value=misafir.get_durum_display()) # Durum sütunu
        ws.cell(row=row_num, column=8, value=misafir.dogum_tarihi.strftime('%d.%m.%Y') if misafir.dogum_tarihi else '-')
        ws.cell(row=row_num, column=9, value=misafir.dogum_yeri if misafir.dogum_yeri else '-')
        ws.cell(row=row_num, column=10, value=misafir.yatak_no.yatak_numarasi if misafir.yatak_no else '-')
        ws.cell(row=row_num, column=11, value=misafir.adres if misafir.adres else '-') # Adres eklendi
       # ws.cell(row=row_num, column=12, value=misafir.get_sosyal_guvence_display() if misafir.sosyal_guvence else '-') # Sosyal Güvence eklendi
        ws.cell(row=row_num, column=12, value=misafir.sosyal_guvence.ad if misafir.sosyal_guvence else '-')
        # ws.cell(row=row_num, column=11, value=kalis_suresi) # Kalış süresi kaldırıldı
        
        # Her hücreye border ekle
        for col_num in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col_num).border = thin_border

        row_num += 1

    # Sütun genişliklerini otomatik ayarla (burada bir değişiklik yok)
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        if adjusted_width > 50:
            adjusted_width = 50
        ws.column_dimensions[column].width = adjusted_width

    # HTTP Response oluştur ve Excel dosyasını ekle
    # Dosya adı sadece liste türünü içerecek, tarih kaldırıldı
    today_str = date.today().strftime("%d-%m-%Y")
    file_name = ""
    if liste_turu == 'aktifler':
        file_name = f"Kalanlar_{today_str}.xlsx" # Sonuç: Kalanlar_14-07-2025.xlsx
    elif liste_turu == 'pasifler':
        file_name = f"Cikanlar_{today_str}.xlsx" # Sonuç: Cikanlar_14-07-2025.xlsx
    else:
        file_name = f"Tum_Kisiler_{today_str}.xlsx" # Sonuç: Tum_Misafirler_14-07-2025.xlsx

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{file_name}"'
    wb.save(response)
    return response

<<<<<<< HEAD
=======
# Diğer view'larınız buraya devam eder...



>>>>>>> 3df35d64b63d79cb98d0843a2f23eefade12dd17
def aylik_analiz(request):
    current_year = date.today().year # Mevcut yıl bilgisi
    aylik_veriler = [] # Her ayın istatistiklerini tutacak liste

    # Her ay için döngü (Ocak'tan Aralık'a)
    for ay in range(1, 13):
        ay_baslangic_date = date(current_year, ay, 1)
        ay_bitis_gunu = monthrange(current_year, ay)[1]
        ay_bitis_date = date(current_year, ay, ay_bitis_gunu)

        ay_baslangic_dt = datetime.combine(ay_baslangic_date, time.min).astimezone(timezone.get_current_timezone())
        ay_bitis_dt = datetime.combine(ay_bitis_date, time.max).astimezone(timezone.get_current_timezone())

        # 1. Ayda Girenler (Yeni Dosya Kayıtları - Sadece Bu Ayda Giriş Yapanlar)
        # Bu, her yeni kaydı (dosya açılışını) ayrı sayar.
        giren_sayisi = Misafir.objects.filter(
            giris_tarihi__year=current_year,
            giris_tarihi__month=ay
        ).count()
        
        # 2. Ayda Çıkanlar (Sadece Bu Ayda Çıkış Yapanlar)
        cikan_sayisi = Misafir.objects.filter(
            cikis_tarihi__year=current_year,
            cikis_tarihi__month=ay
        ).count()
        
        # 3. Ay İçinde Konaklayanlar (Aylık Tablodaki "Kalan" Sütunu)
        # Konaklama süresi o ay ile kesişen herkes.
        ay_ici_konaklayanlar_sayisi = Misafir.objects.filter(
            giris_tarihi__lte=ay_bitis_dt 
        ).filter(
            Q(cikis_tarihi__isnull=True) | Q(cikis_tarihi__gte=ay_baslangic_dt) 
        ).count()
        
        aylik_veriler.append({
            'ay': ay,
            'giren': giren_sayisi,
            'cikan': cikan_sayisi,
            'kalan': ay_ici_konaklayanlar_sayisi, 
            'toplam': giren_sayisi, # Bu ayda açılan dosya/kayıt sayısı
        })

    # --- Genel Toplamlar (Yıl Bazında veya Tüm Zamanlar Bazında) ---

    # Tüm Yıllara Ait Benzersiz Dosya Sayısı (Sisteme kayıtlı toplam benzersiz misafir sayısı)
    # giris_tarihi filtresi yok, tüm Misafir objeleri arasında benzersiz TC kimlik no'ları sayar.
    genel_toplam_tum_zamanlar_benzersiz_dosya_sayisi = Misafir.objects.values('tc_kimlik_no').distinct().count()
    
    # Yıllara Göre Benzersiz Dosya Sayıları (Örnek: {2014: 50, 2015: 70, ...})
    # Her yıl için o yıl ilk kez giriş yapmış benzersiz misafir sayısını bulur.
    yillik_dosya_sayilari = {}
    # Misafirler arasında kayıtlı en eski ve en yeni yılı buluruz
    min_year = Misafir.objects.all().order_by('giris_tarihi').first()
    max_year = Misafir.objects.all().order_by('-giris_tarihi').first()

    if min_year and max_year: # Eğer veritabanında kayıt varsa
        for year in range(min_year.giris_tarihi.year, max_year.giris_tarihi.year + 1):
            # O yıl içinde ilk kez giriş yapmış (benzersiz TC kimlik no'ya göre) misafirler
            count = Misafir.objects.filter(
                giris_tarihi__year=year
            ).values('tc_kimlik_no').distinct().count()
            yillik_dosya_sayilari[year] = count

    # Yıl Boyunca Girenlerin Toplamı (mevcut yıl için aylık girenlerin genel toplamı)
    genel_giren_toplam = sum(item['giren'] for item in aylik_veriler)

    # Yıl Boyunca Çıkanların Toplamı (mevcut yıl için aylık çıkanların genel toplamı)
    genel_cikan_toplam = sum(item['cikan'] for item in aylik_veriler)

    # Mevcut Kalan Kişi Sayısı (Sayfanın yüklendiği an itibarıyla aktif olarak konaklayanlar)
    mevcut_kalan_kisi_sayisi = Misafir.objects.filter(durum='AKTIF').count() 

    # --- Şablona Gönderilecek Veriler ---
    context = {
        'aylik_veriler': aylik_veriler,
        # HTML'deki 'Toplam Dosya Sayısı' kartı için tüm zamanların toplamını gönderiyoruz
        'toplam_kisi': genel_toplam_tum_zamanlar_benzersiz_dosya_sayisi, 
        
        "genel_giren_toplam": genel_giren_toplam, # HTML'deki 'Yıl Boyunca Giren'
        "genel_cikan_toplam": genel_cikan_toplam, # HTML'deki 'Yıl Boyunca Çıkan'
        'toplam_kalan': mevcut_kalan_kisi_sayisi, # HTML'deki 'Mevcut Kalan Kişi'
        
        # 'Yıl Boyunca Toplam Dosya' kartı, aylık girenlerin toplamını (bu yıla ait) gösterir.
        # Eğer bu kartın da tüm zamanlardaki benzersiz dosya sayısını göstermesini isterseniz,
        # yukarıdaki 'toplam_kisi' değişkenini buraya da atayabilirsiniz.
        'genel_aylik_toplam': genel_giren_toplam, 
        
        'yil': current_year, # Sayfa başlığındaki yıl
        'yillik_dosya_sayilari': yillik_dosya_sayilari, # Yeni eklenen yıllara göre dosya sayıları
        'current_section': 'aylik_analiz',
    }
    return render(request, 'bimekan/aylik_analiz.html', context)
    
<<<<<<< HEAD
=======

# Günlük yoklama fonksiyonu (değişiklik yok)
>>>>>>> 3df35d64b63d79cb98d0843a2f23eefade12dd17
@login_required
def gunluk_yoklama(request):
    aktif_kisiler = Misafir.objects.filter(durum='AKTIF')
    yoklama_durumlari = YoklamaDurumu.objects.all()

    if request.method == 'POST':
        form = GunlukYoklamaForm(request.POST, kisiler=aktif_kisiler, durumlar=yoklama_durumlari)
        if form.is_valid():
            eklenen_kayit_sayisi = 0
            for kisi in aktif_kisiler:
                durum = form.cleaned_data.get(f'durum_{kisi.id}')
                if durum:
                    if not YoklamaKaydi.objects.filter(kisi=kisi, tarih=date.today()).exists():
                        YoklamaKaydi.objects.create(kisi=kisi, tarih=date.today(), durum=durum)
                        eklenen_kayit_sayisi += 1

            if eklenen_kayit_sayisi > 0:
                messages.success(request, f"{eklenen_kayit_sayisi} kişi için yoklama kaydedildi.")
            else:
                messages.info(request, "Bugün için zaten yoklama kaydı var. Yeni kayıt eklenmedi.")
            return redirect('gunluk_yoklama')
        else:
            field_map = {f'durum_{kisi.id}': form[f'durum_{kisi.id}'] for kisi in aktif_kisiler}
            
            return render(request, 'Bimekan/gunluk_yoklama.html', {
                'form': form,
                'aktif_kisiler': aktif_kisiler,
                'field_map': field_map,
                'today': date.today(),
                'current_section': 'gunluk-yoklama',
            })
    else:
        form = GunlukYoklamaForm(kisiler=aktif_kisiler, durumlar=yoklama_durumlari)
        kisi_alan_listesi = [(kisi, form[f'durum_{kisi.id}']) for kisi in aktif_kisiler]
        
        return render(request, 'Bimekan/gunluk_yoklama.html', {
            'form': form,
            'aktif_kisiler': aktif_kisiler,
            'kisi_alan_listesi': kisi_alan_listesi,
            'today': date.today(),
            'current_section': 'gunluk-yoklama',
        })

<<<<<<< HEAD
=======

# Yoklama listesi fonksiyonu (değişiklik yok)
>>>>>>> 3df35d64b63d79cb98d0843a2f23eefade12dd17
@login_required
def yoklama_listesi(request):
    arama = request.GET.get('q', '')
    tarih_str = request.GET.get('tarih', '')
    baslangic = request.GET.get('baslangic', '')
    bitis = request.GET.get('bitis', '')

    today = date.today()
    year = today.year
    month = today.month

    seven_days_ago = today - timedelta(days=7)
    ay_baslangici = date(year, month, 1)
    ay_sonu_gunu = monthrange(year, month)[1]
    ay_bitisi = date(year, month, ay_sonu_gunu)

    kayitlar = YoklamaKaydi.objects.select_related('kisi', 'durum')\
        .filter(tarih__range=(ay_baslangici, ay_bitisi))\
        .order_by('-tarih')

    if arama:
        kayitlar = kayitlar.filter(
            Q(kisi__ad__icontains=arama) |
            Q(kisi__soyad__icontains=arama) |
            Q(kisi__tc_kimlik_no__icontains=arama)
        )

    if tarih_str:
        try:
            tarih = datetime.strptime(tarih_str, "%Y-%m-%d").date()
            kayitlar = kayitlar.filter(tarih=tarih)
        except ValueError:
            pass

    elif baslangic and bitis:
        try:
            start = datetime.strptime(baslangic, "%Y-%m-%d").date()
            end = datetime.strptime(bitis, "%Y-%m-%d").date()
            kayitlar = kayitlar.filter(tarih__range=[start, end])
        except ValueError:
            pass

    toplam_var = kayitlar.filter(durum__ad__iexact='var').count()
    toplam_yok = kayitlar.filter(durum__ad__iexact='yok').count()
    toplam_hasta = kayitlar.filter(durum__ad__iexact='hasta').count()
    toplam_izinli = kayitlar.filter(durum__ad__iexact='izinli').count()
    toplam_kayit = kayitlar.count()

    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'

    if is_ajax:
        return render(request, 'yoklama_tablo.html', {
            'kayitlar': kayitlar
        })
    else:
        return render(request, 'bimekan/yoklama_listesi.html', {
            'kayitlar': kayitlar,
            'arama': arama,
            'tarih': tarih_str,
            'baslangic': baslangic,
            'bitis': bitis,
            'today': today.strftime('%Y-%m-%d'),
            'seven_days_ago': seven_days_ago.strftime('%Y-%m-%d'),
            'current_section': 'yoklama-listesi',
            'toplam_var': toplam_var,
            'toplam_yok': toplam_yok,
            'toplam_hasta': toplam_hasta,
            'toplam_izinli': toplam_izinli,
            'toplam_kayit': toplam_kayit,
        })


<<<<<<< HEAD
=======
# Yoklama düzenle fonksiyonu (değişiklik yok)
>>>>>>> 3df35d64b63d79cb98d0843a2f23eefade12dd17
@login_required
def yoklama_duzenle(request, id):
    kayit = get_object_or_404(YoklamaKaydi, id=id)

    if request.method == 'POST':
        durum_id = request.POST.get('durum')
        if durum_id:
            kayit.durum_id = durum_id
            kayit.save()
            messages.success(request, "Yoklama kaydı güncellendi.")
            return redirect('yoklama_listesi')

    durumlar = YoklamaDurumu.objects.all()
    return render(request, 'bimekan/yoklama_duzenle.html', {
        'kayit': kayit,
        'durumlar': durumlar
    })

<<<<<<< HEAD
=======
# Yoklama sil fonksiyonu (değişiklik yok)
>>>>>>> 3df35d64b63d79cb98d0843a2f23eefade12dd17
@login_required
def yoklama_sil(request, id):
    kayit = get_object_or_404(YoklamaKaydi, id=id)
    kayit.delete()
    messages.success(request, "Yoklama kaydı silindi.")
    return redirect('yoklama_listesi')

<<<<<<< HEAD
=======
# Yoklama AJAX fonksiyonu (değişiklik yok)
>>>>>>> 3df35d64b63d79cb98d0843a2f23eefade12dd17
@login_required
def yoklama_ajax(request):
    arama = request.GET.get('q', '')
    tarih_str = request.GET.get('tarih', '')

    kayitlar = YoklamaKaydi.objects.select_related('kisi', 'durum')

    if arama:
        kayitlar = kayitlar.filter(
            Q(kisi__ad__icontains=arama) |
            Q(kisi__soyad__icontains=arama) |
            Q(kisi__tc_kimlik_no__icontains=arama)
        )

    if tarih_str:
        try:
            tarih = datetime.strptime(tarih_str, "%Y-%m-%d").date()
            kayitlar = kayitlar.filter(tarih=tarih)
        except ValueError:
            pass

    html = render_to_string('bimekan/yoklama_tablo.html', {'kayitlar': kayitlar})
    return HttpResponse(html)

def tc_kontrol_view(request):
    tc = request.GET.get('tc')
    misafir = Misafir.objects.filter(tc_kimlik_no=tc).first()
    if misafir:
        return JsonResponse({'status': 'var', 'misafir_id': misafir.id})
    else:
        return JsonResponse({'status': 'yok'})

<<<<<<< HEAD
=======
# Register View (değişiklik yok)
>>>>>>> 3df35d64b63d79cb98d0843a2f23eefade12dd17
class RegisterView(generic.CreateView):
    form_class = UserCreationForm
    success_url = reverse_lazy('login')
    template_name = 'registration/register.html'

    def form_valid(self, form):
        messages.success(self.request, "Hesabınız başarıyla oluşturuldu! Lütfen giriş yapın.")
        return super().form_valid(form)

    def form_invalid(self, form):
        for field, errors in form.errors.items():
            for error in errors:
                label = form.fields[field].label if field in form.fields else "Genel Hata"
                messages.error(self.request, f"'{label}' alanında hata: {error}")
        return super().form_invalid(form)

<<<<<<< HEAD
=======

# Misafir kayıt ve giriş view (değişiklik yok)
>>>>>>> 3df35d64b63d79cb98d0843a2f23eefade12dd17
@login_required
def misafir_kayit_ve_giris(request):
    """
    Yeni bir misafiri sisteme kaydeder ve otomatik olarak giriş işlemini oluşturur.
    GET isteğinde boş form gösterilir, POST isteğinde form verileri işlenir.
    """

    if request.method == 'POST':
        form = MisafirKayitForm(request.POST, request.FILES)
        if form.is_valid():
            misafir = form.save(commit=False)
            misafir.durum = 'AKTIF'
            misafir.save()

            if misafir.yatak_no:
                misafir.yatak_no.dolu_mu = True
                misafir.yatak_no.save()

            try:
                giris_islem_turu = IslemTuru.objects.get(ad__iexact='Giriş')
            except IslemTuru.DoesNotExist:
                messages.error(request, "Hata: 'Giriş' işlem türü bulunamadı.")
                if misafir.yatak_no:
                    misafir.yatak_no.dolu_mu = False
                    misafir.yatak_no.save()
                misafir.delete()
                return redirect('misafir_kayit_ve_giris')

            Islem.objects.create(
                misafir=misafir,
                islem_turu=giris_islem_turu,
            )

            messages.success(request, f"'{misafir.ad} {misafir.soyad}' başarıyla kaydedildi ve giriş işlemi yapıldı.")
            return redirect('misafir_detay', pk=misafir.pk)
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    label = form.fields[field].label if field in form.fields else "Genel Hata"
                    messages.error(request, f"'{label}' alanında hata: {error}")
    else:
        form = MisafirKayitForm()

    context = {
        'form': form,
        'title': 'Yeni Misafir Kayıt & Giriş',
        'current_section': 'new-misafir-form',
    }
    return render(request, 'bimekan/misafir_kayit_giris.html', context)

<<<<<<< HEAD
=======
import math

# Misafir listeleri view (değişiklik yok)
>>>>>>> 3df35d64b63d79cb98d0843a2f23eefade12dd17
@login_required
def misafir_listeleri(request, liste_turu='tumu'):
    base_queryset = Misafir.objects.all()
    title = "Tüm Kişiler Listesi"
    current_queryset = base_queryset.order_by('ad', 'soyad')

    if liste_turu == 'aktifler':
        current_queryset = base_queryset.filter(durum='AKTIF').order_by('ad', 'soyad')
        title = "Aktif Kişiler Listesi"
    elif liste_turu == 'pasifler':
        current_queryset = base_queryset.filter(durum='PASIF').order_by('ad', 'soyad')
        title = "Pasif Kişiler Listesi"

    query = request.GET.get('q')
    if query:
        # Sorguyu boşluklara göre kelimelere ayırıyoruz.
        # Örneğin "Cem Edman" -> ['Cem', 'Edman']
        arama_kelimeleri = query.split()
        
        # Ana Q objesi, her kelimenin kendi içinde OR'lanmış alanlarda bulunmasını sağlayacak.
        # Örneğin 'Cem' kelimesi ad VEYA soyadda olsun.
        # 'Edman' kelimesi ad VEYA soyadda olsun.
        
        # Sonra bu her bir kelime için oluşturulan Q objelerini AND ile birleştireceğiz.
        # Yani 'Cem' (adı/soyadı içinde) VE 'Edman' (adı/soyadı içinde)
        
        # İlk olarak, kelimelerin ad/soyad/TC/dosya_no alanlarında bulunması için genel bir filtre oluşturalım.
        # Bu kısım, kullanıcı tek kelime aradığında da doğru çalışır.
        overall_q_object = Q()
        for kelime in arama_kelimeleri:
            # Her kelimenin ad, soyad, TC veya dosya no'da geçmesi YETERLİ (OR mantığı)
            overall_q_object &= (
                Q(ad__icontains=kelime) |
                Q(soyad__icontains=kelime) |
                Q(tc_kimlik_no__icontains=kelime) |
                Q(dosya_no__icontains=kelime)
            )

        # Eğer arama kelime sayısı 2 veya daha fazlaysa ve kullanıcı ad soyad arıyor gibiyse,
        # daha spesifik "Adında X VE Soyadında Y" mantığını da deneyelim.
        if len(arama_kelimeleri) >= 2:
            # Ad ve Soyad kombinasyonu için ayrı bir Q objesi oluşturuyoruz.
            # İlk kelime ad'da OLSUN VE İkinci kelime soyadda OLSUN
            # VEYA İkinci kelime ad'da OLSUN VE Birinci kelime soyadda OLSUN (tersine arama)
            name_surname_q_object = Q()
            
            # İlk kelimeyi ad, ikinci kelimeyi soyad olarak eşleştirmeye çalış.
            # Örneğin: query = "Cem Edman" -> kelimeler = ['Cem', 'Edman']
            # Q(ad__icontains='Cem') & Q(soyad__icontains='Edman')
            name_surname_q_object |= (
                Q(ad__icontains=arama_kelimeleri[0]) & 
                Q(soyad__icontains=arama_kelimeleri[1])
            )
            
            # Tersini de kontrol et, belki "Edman Cem" yazmıştır.
            # Q(ad__icontains='Edman') & Q(soyad__icontains='Cem')
            if len(arama_kelimeleri) >= 2: # arama kelime sayısı en az 2 ise bu kontrolü yap
                name_surname_q_object |= (
                    Q(ad__icontains=arama_kelimeleri[1]) & 
                    Q(soyad__icontains=arama_kelimeleri[0])
                )
            
            # Eğer 3 veya daha fazla kelime varsa, geri kalan kelimeleri genel filtreye ekle.
            # Örneğin "Ayşe Sultan Yılmaz" -> Ayşe (ad/soyad) VE Sultan (ad/soyad) VE Yılmaz (ad/soyad)
            # Bu, overall_q_object zaten ele alıyor.
            
            # Sonuç olarak, genel kelime bazlı filtre VE ad-soyad kombinasyon filtrelerinin OR'unu alıyoruz.
            # Bu, her iki tür aramayı da destekler.
            final_q_object = overall_q_object | name_surname_q_object
            
        else: # Tek kelime arama
            final_q_object = overall_q_object


        current_queryset = current_queryset.filter(final_q_object).order_by('ad', 'soyad')

        if not current_queryset.exists():
            messages.info(request, f"'{query}' için aradığınız kriterlere uygun kişi bulunamadı.")

    paginator = Paginator(current_queryset, 20)
    page_number = request.GET.get('page')
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

# # # # # ÖNEMLİ DEĞİŞİKLİK BURADA BAŞLIYOR # # # # # 
    # Yatak no listede Nan Geliyordu o nedenle koydum bu kodu 
    # Şablona göndermeden önce her misafir objesinin yatak_numarasi'nı kontrol ediyoruz.
    processed_misafirler = []
    for misafir in page_obj.object_list:
        # # # # # BURADAKİ LOG KODUNU EKLEYİN # # # # #
        if misafir.yatak_no: # Eğer yatak ilişkisi varsa
            yatak_numarasi_degeri = misafir.yatak_no.yatak_numarasi
            # Bu değeri ve tipini terminale yazdıracak
            
            
            # Eski kontrol mantığımız
            if isinstance(yatak_numarasi_degeri, (int, float)) and math.isnan(yatak_numarasi_degeri):
                misafir.display_yatak_no = "Yok" 
            elif yatak_numarasi_degeri is None or str(yatak_numarasi_degeri).strip() == "" or str(yatak_numarasi_degeri).strip().upper() == "NAN":
                misafir.display_yatak_no = "Yok" 
            else:
                misafir.display_yatak_no = str(yatak_numarasi_degeri)
        else:
            misafir.display_yatak_no = "-"
        # # # # # LOG KODU BURADA BİTİYOR # # # # #
        processed_misafirler.append(misafir)
    # # # # # ÖNEMLİ DEĞİŞİKLİK BURADA BİTİYOR # # # # #

    toplam_kisi_sayisi = base_queryset.count()
    aktif_sayisi = base_queryset.filter(durum='AKTIF').count()
    pasif_sayisi = base_queryset.filter(durum='PASIF').count()

    context = {
        'misafirler': page_obj.object_list,
        'page_obj': page_obj,
        'query': query,
        'title': title,
        'liste_turu': liste_turu,
        'current_section': f'{liste_turu}-misafirler-list',
        'now_dt': timezone.now(),
        'toplam_kisi_sayisi': toplam_kisi_sayisi,
        'aktif_sayisi': aktif_sayisi,
        'pasif_sayisi': pasif_sayisi,
    }

    return render(request, 'bimekan/misafir_listesi.html', context)

<<<<<<< HEAD
=======

# Misafir detay view (değişiklik yok)
>>>>>>> 3df35d64b63d79cb98d0843a2f23eefade12dd17
@login_required
def misafir_detay(request, pk):
    misafir = get_object_or_404(Misafir, pk=pk)
    islemler = Islem.objects.filter(misafir=misafir).order_by('-islem_zamani')

    islem_paginator = Paginator(islemler, 20)
    islem_page_number = request.GET.get('islem_page')
    try:
        islem_page_obj = islem_paginator.page(islem_page_number)
    except PageNotAnInteger:
        islem_page_obj = islem_paginator.page(1)
    except EmptyPage:
        islem_page_obj = islem_paginator.page(islem_paginator.num_pages)

    # 🔹 Kalış Süresi Hesabı
    toplam_kalis_suresi = None
    if misafir.durum == 'AKTIF':
        if misafir.giris_tarihi:
            if timezone.is_aware(misafir.giris_tarihi):
                delta = timezone.now() - misafir.giris_tarihi
            else:
                delta = datetime.now() - misafir.giris_tarihi 
            toplam_kalis_suresi = f"{delta.days} gün"
    elif misafir.giris_tarihi and misafir.cikis_tarihi:
        delta = misafir.cikis_tarihi - misafir.giris_tarihi
        toplam_kalis_suresi = f"{delta.days} gün"

    # 🔹 İşlem Tutarı Hesabı
    toplam_islem_tutari = islemler.aggregate(Sum('tutar'))['tutar__sum']
    if toplam_islem_tutari is None:
        toplam_islem_tutari = 0.00

    # 🔹 Yaş Hesabı
    yas = None
    if misafir.dogum_tarihi:
        today = date.today()
        yas = today.year - misafir.dogum_tarihi.year
        if (today.month, today.day) < (misafir.dogum_tarihi.month, misafir.dogum_tarihi.day):
            yas -= 1

    # 🔹 Sayfa Verileri
    context = {
        'title': f"{misafir.ad} {misafir.soyad} Detayları",
        'misafir': misafir,
        'islemler': islem_page_obj.object_list,
        'islem_page_obj': islem_page_obj,
        'toplam_kalis_suresi': toplam_kalis_suresi,
        'toplam_islem_tutari': toplam_islem_tutari,
        'yas': yas,
        'current_section': 'misafir-detay-form',
    }
    return render(request, 'bimekan/misafir_detay.html', context)

from .models import GiyimIslem

<<<<<<< HEAD
=======
# View to perform a transaction for an existing guest
>>>>>>> 3df35d64b63d79cb98d0843a2f23eefade12dd17
@login_required
def misafir_islem_yap(request, pk):
    misafir = get_object_or_404(Misafir, pk=pk)

    if request.method == 'POST':
        form = MisafirIslemForm(request.POST, instance=Islem())
        if form.is_valid():
            islem = form.save(commit=False)
            islem.misafir = misafir

            islem_turu_adi_lower = islem.islem_turu.ad.lower()

<<<<<<< HEAD
            # ❗ ESKİ AYNİ YARDIM BLOĞU BURADAN KALDIRILDI ❗
            # Stok güncelleme ve GiyimIslem oluşturma mantığı artık Islem modelinin save metodunda.
            # Dolayısıyla burada sadece islem.aciklama ve islem.tutar ayarı kalabilir.
            if islem_turu_adi_lower == 'ayni yardım': # 'ayni' yerine 'ayni yardım' olarak güncellendi
                urun = form.cleaned_data.get('urun')
                miktar = form.cleaned_data.get('miktar')
                if urun and miktar:
                    islem.aciklama = f"Ayni yardım: {miktar} adet {urun.ad} verildi. " + (islem.aciklama or "")
                    islem.tutar = 0 # Ayni yardımların tutarı 0 olmalı

            # 🔹 Giriş kontrolü (aynı kaldı)
=======
            # 🔹 Ayni Yardım İşlemi
            if islem_turu_adi_lower == 'ayni':
                urun = form.cleaned_data.get('urun')
                miktar = form.cleaned_data.get('miktar')

                if urun and miktar:
                    urun.mevcut_adet = max(urun.mevcut_adet - miktar, 0)
                    urun.save()

                    GiyimIslem.objects.create(
                        urun=urun,
                        miktar=miktar,
                        alici=misafir,
                        islem_turu="Çıkış",
                        aciklama=f"Ayni yardım - {miktar} adet {urun.ad} verildi"
                    )

                    islem.aciklama = f"{miktar} adet {urun.ad} verildi"
                    islem.tutar = 0

            # 🔹 Giriş kontrolü
>>>>>>> 3df35d64b63d79cb98d0843a2f23eefade12dd17
            if islem_turu_adi_lower == 'giriş' and misafir.durum == 'AKTIF':
                messages.error(request, f"'{misafir.ad} {misafir.soyad}' zaten aktif durumda. Tekrar giriş yapılamaz.")
                return redirect('misafir_detay', pk=misafir.pk)

<<<<<<< HEAD
            # 🔹 Çıkış kontrolü (aynı kaldı)
=======
            # 🔹 Çıkış kontrolü
>>>>>>> 3df35d64b63d79cb98d0843a2f23eefade12dd17
            if islem_turu_adi_lower == 'çıkış' and misafir.durum == 'PASIF':
                messages.error(request, f"'{misafir.ad} {misafir.soyad}' zaten pasif durumda. Tekrar çıkış yapılamaz.")
                return redirect('misafir_detay', pk=misafir.pk)

<<<<<<< HEAD
            # 🔹 Giriş tarihi eşleştirme (aynı kaldı, Misafir modeli güncellendi)
            if islem_turu_adi_lower == 'giriş':
                giris_tarihi = form.cleaned_data.get('giris_tarihi')
                if giris_tarihi:
                    islem.giris_tarihi = giris_tarihi # İşlem kaydına giriş tarihini de ekleyebiliriz (opsiyonel)
                    misafir.giris_tarihi = giris_tarihi
                    misafir.durum = 'AKTIF' # Girişte durumu aktif yap

            elif islem_turu_adi_lower == 'çıkış':
                cikis_tarihi = form.cleaned_data.get('cikis_tarihi')
                cikis_nedeni = form.cleaned_data.get('cikis_nedeni')
                if cikis_tarihi:
                    islem.cikis_tarihi = cikis_tarihi # İşlem kaydına çıkış tarihini de ekleyebiliriz (opsiyonel)
                    misafir.cikis_tarihi = cikis_tarihi
                if cikis_nedeni:
                    misafir.cikis_nedeni = cikis_nedeni
                
                # Misafir durumu ve yatak boşaltma
                misafir.durum = 'PASIF'
                if misafir.yatak_no:
                    yatak = misafir.yatak_no
                    yatak.dolu_mu = False
                    yatak.save()
                    misafir.yatak_no = None # Misafirin yatağını boşalt
                    messages.info(request, f"'{misafir.ad} {misafir.soyad}' için yatak numarası boşaltıldı.")
                else:
                    messages.warning(request, f"'{misafir.ad} {misafir.soyad}' için atanmış yatak bulunamadı.")
                messages.info(request, f"'{misafir.ad} {misafir.soyad}' için '{islem.islem_turu.ad}' işlemi başarıyla yapıldı. Yeni durum: '{misafir.get_durum_display()}'.")
            else: # Diğer işlem türleri (nakdi yardım, vb.)
                 messages.success(request, f"'{misafir.ad} {misafir.soyad}' için '{islem.islem_turu.ad}' işlemi başarıyla kaydedildi.")

            misafir.save() # Misafir objesini kaydet (durum ve tarih güncellemeleri için)
            islem.save() # Islem objesini kaydet (bu, modeldeki save metodunu tetikler ve stoğu günceller)
            
=======
            # 🔹 Giriş tarihi eşleştirme
            if islem_turu_adi_lower == 'giriş':
                giris_tarihi = form.cleaned_data.get('giris_tarihi')
                if giris_tarihi:
                    islem.giris_tarihi = giris_tarihi
                    misafir.giris_tarihi = giris_tarihi
            elif islem_turu_adi_lower == 'çıkış':
                cikis_tarihi = form.cleaned_data.get('cikis_tarihi')
                if cikis_tarihi:
                    islem.cikis_tarihi = cikis_tarihi

            # 🔹 Durum değişikliği
            if islem.islem_turu.durum_degistirir_mi and islem.islem_turu.yeni_durum:
                misafir.durum = islem.islem_turu.yeni_durum

                if islem_turu_adi_lower == 'çıkış':
                    misafir.cikis_tarihi = form.cleaned_data.get('cikis_tarihi') or timezone.now()
                    misafir.cikis_nedeni = form.cleaned_data.get('cikis_nedeni') or None
                    if misafir.yatak_no:
                        yatak = misafir.yatak_no
                        yatak.dolu_mu = False
                        yatak.save()
                        misafir.yatak_no = None
                        messages.info(request, f"'{misafir.ad} {misafir.soyad}' için yatak numarası boşaltıldı.")
                    else:
                        messages.warning(request, f"'{misafir.ad} {misafir.soyad}' için atanmış yatak bulunamadı.")

                messages.info(request, f"'{misafir.ad} {misafir.soyad}' için '{islem.islem_turu.ad}' işlemi başarıyla yapıldı. Yeni durum: '{misafir.get_durum_display()}'.")
            else:
                messages.success(request, f"'{misafir.ad} {misafir.soyad}' için '{islem.islem_turu.ad}' işlemi başarıyla kaydedildi.")

            misafir.save()
            islem.save()
>>>>>>> 3df35d64b63d79cb98d0843a2f23eefade12dd17
            return redirect('misafir_detay', pk=misafir.pk)

        else:
            for field, errors in form.errors.items():
                for error in errors:
                    label = form.fields[field].label if field in form.fields else "Genel Hata"
                    messages.error(request, f"'{label}' alanında hata: {error}")
    else:
        form = MisafirIslemForm()

    context = {
        'form': form,
        'misafir': misafir,
        'title': f"{misafir.ad} {misafir.soyad} İçin İşlem Yap",
        'current_section': 'misafir-islem-form',
    }
    return render(request, 'bimekan/misafir_islem.html', context)


# Misafir işlem seçim view (değişiklik yok)
@login_required
def misafir_islem_secim(request):
    query = request.GET.get('q')
    misafirler = Misafir.objects.all().order_by('ad', 'soyad')

    if query:
        misafirler = misafirler.filter(
            Q(ad__icontains=query) |
            Q(soyad__icontains=query) |
            Q(tc_kimlik_no__icontains=query) |
            Q(dosya_no__icontains=query)
        )
        if not misafirler.exists():
            messages.info(request, f"Aradığınız kriterlere uygun kişi bulunamadı.")
    
    paginator = Paginator(misafirler, 20)
    page_number = request.GET.get('page')
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    context = {
        'query': query,
        'misafirler': page_obj.object_list,
        'page_obj': page_obj,
        'title': 'İşlem Yapılacak Kişi Seçimi',
        'current_section': 'misafir-islem-secim-form',
    }
    return render(request, 'bimekan/misafir_islem_secim.html', context)




# Misafir güncelleme Class-Based View (değişiklik yok)
class MisafirGuncelleView(LoginRequiredMixin, UpdateView):
    model = Misafir
    form_class = MisafirGuncelleForm
    template_name = 'bimekan/misafir_duzenle.html'
    context_object_name = 'misafir'

    def get_success_url(self):
        messages.success(self.request, f"'{self.object.ad} {self.object.soyad}' bilgileri başarıyla güncellendi.")
        return reverse_lazy('misafir_detay', kwargs={'pk': self.object.pk})
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f"{self.object.ad} {self.object.soyad} - Bilgileri Düzenle"
        context['current_section'] = 'misafir-duzenle-form'
        return context

    def form_valid(self, form):
        if form.cleaned_data['yatak_no'] != self.object.yatak_no:
            if self.object.yatak_no:
                self.object.yatak_no.dolu_mu = False
                self.object.yatak_no.save()
            if form.cleaned_data['yatak_no']:
                form.cleaned_data['yatak_no'].dolu_mu = True
                form.cleaned_data['yatak_no'].save()

        if 'fotograf-clear' in self.request.POST and self.request.POST['fotograf-clear'] == 'on':
            form.instance.fotograf = None
        elif 'fotograf' in self.request.FILES:
            form.instance.fotograf = self.request.FILES['fotograf']

        return super().form_valid(form)

    def form_invalid(self, form):
        for field, errors in form.errors.items():
            for error in errors:
                label = form.fields[field].label if field in form.fields else "Genel Hata"
                messages.error(self.request, f"'{label}' alanında hata: {error}")
        return self.render_to_response(self.get_context_data(form=form))


# Misafir silme Class-Based View (değişiklik yok)
class MisafirSilView(LoginRequiredMixin, DeleteView):
    model = Misafir
    template_name = 'bimekan/misafir_sil_onay.html'
    success_url = reverse_lazy('misafir_listesi', kwargs={'liste_turu': 'tumu'})

    def form_valid(self, form):
        if self.object.yatak_no and self.object.yatak_no.dolu_mu:
            self.object.yatak_no.dolu_mu = False
            self.object.yatak_no.save()
        messages.warning(self.request, f"'{self.object.ad} {self.object.soyad}' kaydı başarıyla silindi.")
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f"'{self.object.ad} {self.object.soyad}' Kaydını Sil"
        return context

# İşlem silme Class-Based View (değişiklik yok)
class IslemSilView(LoginRequiredMixin, DeleteView):
    model = Islem
    template_name = 'bimekan/islem_sil_onay.html'

    def get_success_url(self):
        misafir_pk = self.object.misafir.pk
        messages.warning(self.request, f"'{self.object.misafir.ad} {self.object.misafir.soyad}' için '{self.object.islem_turu.ad}' işlemi başarıyla silindi.")
        return reverse_lazy('misafir_detay', kwargs={'pk': misafir_pk})
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f"'{self.object.islem_turu.ad}' İşlemini Sil"
        return context

# Başvuru formu çıktı view
@login_required
def basvuru_formu_cikti_view(request, pk):
    misafir = get_object_or_404(Misafir, pk=pk)
    detay_url = reverse('misafir_detay', args=[misafir.id])
    context = {
        'misafir': misafir,
        'detay_url': detay_url,
    }
    return render(request, 'bimekan/basvuru_formu.html', context)

from django.urls import reverse
# Açık rıza formu view
@login_required
def acikriza_formu_view(request, pk):
    misafir = get_object_or_404(Misafir, pk=pk)
    detay_url = reverse('misafir_detay', args=[misafir.id])
    context = {
        'misafir': misafir,
        'detay_url': detay_url,
    }
    return render(request, 'bimekan/acikriza_formu.html', context)

# Açık rıza SOYBIS formu view
@login_required
def acikriza_soybis_view(request, pk):
    misafir = get_object_or_404(Misafir, pk=pk)
    detay_url = reverse('misafir_detay', args=[misafir.id])  # 🔹 Geri butonunun çalışması için
    context = {
        'tc_kimlik_no': misafir.tc_kimlik_no,
        'ad_soyad': misafir.ad_soyad,
        'tarih': misafir.giris_tarihi.strftime('%d.%m.%Y') if misafir.giris_tarihi else '',
        'misafir': misafir,               # Eğer template'te .id gibi erişim yapacaksan
        'detay_url': detay_url            # 🔹 HTML'de geri butonuna verilecek URL
    }
    return render(request, 'bimekan/acikriza_soybis.html', context)

from .forms import YatakForm
from collections import defaultdict

def sort_by_oda_no(yatak):
    try:
        oda_no = int(yatak.yatak_numarasi.split('-')[2])
        # return -oda_no   Büyükten küçüğe sıralamak için negatif
        return oda_no  # Küçükten Büyüğpe sıralamak için negatif
    except:
        return 0  # Hata varsa en alta koy

<<<<<<< HEAD
=======

>>>>>>> 3df35d64b63d79cb98d0843a2f23eefade12dd17
def yatak_ekle(request):
    form = YatakForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('yatak_ekle')

    # Tüm yatakları al
    yataklar = Yatak.objects.all()
    for yatak in yataklar:
        yatak.aktif_misafirler = Misafir.objects.filter(yatak_no=yatak, durum='AKTIF')

    # Katlara göre ayır
    kat_1_k, kat_1_b = [], []
    kat_2_k, kat_2_b = [], []
    kat_3_k, kat_3_b = [], []
    kat_teras = []

    for yatak in yataklar:
        kod = yatak.yatak_numarasi.split('-')
        if len(kod) != 3:
            continue  # Geçersiz kod varsa atla

        kat, tip, _ = kod
        # # # # # BURADA DEĞİŞİKLİK YAPILMALI # # # # #
        tip_upper = tip.upper() # Tip bilgisini büyük harfe çevirelim

        if kat == '1' and tip_upper != 'T': # 1. kat, Teras olmayanlar
            (kat_1_k if tip_upper == 'K' else kat_1_b).append(yatak)
        elif kat == '2' and tip_upper != 'T': # 2. kat, Teras olmayanlar
            (kat_2_k if tip_upper == 'K' else kat_2_b).append(yatak)
        elif kat == '3' and tip_upper != 'T': # 3. kat, Teras olmayanlar
            (kat_3_k if tip_upper == 'K' else kat_3_b).append(yatak)
        
        # Tip bilgisi 'T' ise teras olarak kabul et
        elif tip_upper == 'T':
            kat_teras.append(yatak)
        # else:
            # logger.warning(f"Bilinmeyen kat/tip kombinasyonu atlandı: {yatak.yatak_numarasi}")
        # # # # # DEĞİŞİKLİK BİTİŞİ # # # # #

    # Oda numarasına göre küçükten büyüğe sırala
    def sort_by_oda_no(yatak):
        try:
            return int(yatak.yatak_numarasi.split('-')[2])
        except:
            return 0

    kat_1_k = sorted(kat_1_k, key=sort_by_oda_no)
    kat_1_b = sorted(kat_1_b, key=sort_by_oda_no)
    kat_2_k = sorted(kat_2_k, key=sort_by_oda_no)
    kat_2_b = sorted(kat_2_b, key=sort_by_oda_no)
    kat_3_k = sorted(kat_3_k, key=sort_by_oda_no)
    kat_3_b = sorted(kat_3_b, key=sort_by_oda_no)
    kat_teras = sorted(kat_teras, key=sort_by_oda_no)

    return render(request, 'bimekan/yatak_ekle.html', {
        'form': form,
        'kat_1_k': kat_1_k,
        'kat_1_b': kat_1_b,
        'kat_2_k': kat_2_k,
        'kat_2_b': kat_2_b,
        'kat_3_k': kat_3_k,
        'kat_3_b': kat_3_b,
        'kat_teras': kat_teras,
    })

<<<<<<< HEAD
=======

>>>>>>> 3df35d64b63d79cb98d0843a2f23eefade12dd17
def yatak_sil(request, pk):
    yatak = get_object_or_404(Yatak, pk=pk)

    # 🔐 Doğru filtreleme: aktif misafir var mı?
    if Misafir.objects.filter(yatak_no=yatak, durum='AKTIF').exists():
        messages.error(request, f"{yatak.yatak_numarasi} numaralı yatakta aktif misafir bulunduğu için silme engellendi.")
        return redirect('yatak_ekle')

    yatak.delete()
    messages.success(request, f"{yatak.yatak_numarasi} numaralı yatak başarıyla silindi.")
    return redirect('yatak_ekle')

<<<<<<< HEAD
=======

>>>>>>> 3df35d64b63d79cb98d0843a2f23eefade12dd17
def yatak_duzenle(request, pk):
    yatak = get_object_or_404(Yatak, pk=pk)

    # 🔍 Bu yatak aktif bir misafire atanmış mı?
    aktif_misafir_var = Misafir.objects.filter(yatak=yatak, aktif=True).exists()

    if aktif_misafir_var:
        messages.error(
            request,
            f"{yatak.yatak_numarasi} numaralı yatak aktif bir misafire atanmış durumda. Düzenleme yapılamaz."
        )
        return redirect('yatak_ekle')

    form = YatakForm(request.POST or None, instance=yatak)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, f"{yatak.yatak_numarasi} numaralı yatak güncellendi.")
        return redirect('yatak_ekle')

    # Normalde inline çalıştığı için GET ile bu sayfa render edilmiyor
    return redirect('yatak_ekle')

<<<<<<< HEAD
=======



>>>>>>> 3df35d64b63d79cb98d0843a2f23eefade12dd17
def sosyal_guvence_ekle(request):
    form = SosyalGuvenceForm(request.POST or None)
    if request.method == 'POST':
        if 'duzenle_id' in request.POST:
            guvence = get_object_or_404(SosyalGuvence, pk=request.POST['duzenle_id'])
            duzenle_form = SosyalGuvenceForm(request.POST, instance=guvence)
            if duzenle_form.is_valid():
                duzenle_form.save()
                return redirect('sosyal_guvence_ekle')
        elif form.is_valid():
            form.save()
            return redirect('sosyal_guvence_ekle')

    guvenceler = SosyalGuvence.objects.all().order_by('ad')
    return render(request, 'bimekan/sosyal_guvence_ekle.html', {
        'form': form,
        'guvenceler': guvenceler,
        'current_section': 'sosyal_guvence',
    })

def sosyal_guvence_sil(request, pk):
    guvence = get_object_or_404(SosyalGuvence, pk=pk)
    if request.method == 'POST':
        guvence.delete()
        return redirect('sosyal_guvence_ekle')
    
<<<<<<< HEAD
=======

>>>>>>> 3df35d64b63d79cb98d0843a2f23eefade12dd17
def islem_detay(request):
    islem_turu = request.GET.get("islem_turu")
    kurum = request.GET.get("kurum")
    tarih1 = request.GET.get("tarih1")
    tarih2 = request.GET.get("tarih2")
    
    
    bugun = date.today()
    ay_basi = bugun.replace(day=1)

    filtre = Q(islem_zamani__date__gte=ay_basi)  # İlk yüklendiğinde bu ay

    if tarih1 and tarih2:
        filtre &= Q(islem_zamani__date__range=[tarih1, tarih2])
    elif tarih1:
        filtre &= Q(islem_zamani__date__gte=tarih1)
    elif tarih2:
        filtre &= Q(islem_zamani__date__lte=tarih2)

    if islem_turu:
        filtre &= Q(islem_turu_id=islem_turu)
    
    if kurum:
        filtre &= Q(kurum_id=kurum)

    islemler = Islem.objects.filter(filtre).order_by("-islem_zamani")
    
    toplam_tutar = islemler.aggregate(Sum("tutar"))["tutar__sum"] or 0
    context = {
        "islemler": islemler,
        "islem_turleri": IslemTuru.objects.all(),
        "kurumlar": Kurum.objects.all(),
        "toplam_tutar": toplam_tutar,
        "current_section": "islem_detay"
    }
    return render(request, "bimekan/islem_detay.html", context)