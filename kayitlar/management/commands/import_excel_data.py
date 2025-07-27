# kayitlar/management/commands/import_excel_data.py
import openpyxl
from django.core.management.base import BaseCommand, CommandError
from kayitlar.models import Misafir, Yatak, SosyalGuvence # Kendi modellerinizi import ettiğinizden emin olun
from django.utils import timezone
from datetime import datetime, date
import re
import sys

class Command(BaseCommand):
    help = 'Belirtilen Excel dosyasındaki misafir verilerini veritabanına aktarır.'

    def add_arguments(self, parser):
        parser.add_argument('excel_file', type=str, help='İçe aktarılacak Excel dosyasının yolu')
        parser.add_argument(
            '--sheet',
            type=str,
            default=None,
            help='Verilerin bulunduğu Excel sayfasının adı. Belirtilmezse ilk sayfa kullanılır.'
        )

    def handle(self, *args, **kwargs):
        file_path = kwargs['excel_file']
        sheet_name = kwargs['sheet']
        
        try:
            workbook = openpyxl.load_workbook(file_path)
        except FileNotFoundError:
            raise CommandError(f"Hata: '{file_path}' dosyası bulunamadı.")
        except Exception as e:
            raise CommandError(f"Excel dosyasını yüklerken bir hata oluştu: {e}")

        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise CommandError(f"Hata: '{sheet_name}' adında bir sayfa bulunamadı. Mevcut sayfalar: {', '.join(workbook.sheetnames)}")
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active # İlk aktif sayfayı kullan

        self.stdout.write(self.style.SUCCESS(f"'{sheet.title}' sayfasından veri okunuyor..."))

        # Başlık satırını atlamak için (ilk satırın başlık olduğunu varsayıyoruz)
        header = [cell.value for cell in sheet[1]]
        data_rows = list(sheet.iter_rows(min_row=2, values_only=True))

        imported_count = 0
        updated_count = 0
        skipped_count = 0
        
        # Excel sütunlarını Misafir model alanlarına manuel eşleme (Excel sütun başlıklarınıza göre ayarlayın)
        # Örneğin: Eğer Excel'de "T.C." başlıklı bir sütun varsa, onu 'tc_kimlik_no'ya eşleyin.
        column_mapping = {
            'Dosya No': 'dosya_no',
            'Adı': 'ad',
            'Soyadı': 'soyad',
            'T.C. Kimlik No': 'tc_kimlik_no',
            'Giriş Tarihi': 'giris_tarihi',
            'Çıkış Tarihi': 'cikis_tarihi',
            'Doğum Tarihi': 'dogum_tarihi',
            'Doğum Yeri': 'dogum_yeri',
            'Telefonu': 'telefon',
            'Durum Bilgisi': 'durum',
            'Adresi': 'adres',
            'Beyanı': 'beyan',
            'Yatak Numarası': 'yatak_no', # Eğer Excel'de yatak no varsa
            'Sosyal Güvencesi': 'sosyal_guvence', # Eğer Excel'de sosyal güvence varsa
            # 'Fotograf': 'fotograf', # Dosya yüklemesi daha karmaşık olabilir, bu örnekte dahil değil
        }

        # Başlıkları indekslere eşle
        header_indices = {h: idx for idx, h in enumerate(header) if h in column_mapping}

        for row_idx, row_data in enumerate(data_rows, start=2): # Satır numarasını 2'den başlat
            try:
                # Minimum gerekli alanları kontrol edin
                tc_kimlik_no = str(row_data[header_indices.get('T.C. Kimlik No', -1)]).strip() if 'T.C. Kimlik No' in header_indices else None
                ad = str(row_data[header_indices.get('Adı', -1)]).strip() if 'Adı' in header_indices else None
                soyad = str(row_data[header_indices.get('Soyadı', -1)]).strip() if 'Soyadı' in header_indices else None
                giris_tarihi_raw = row_data[header_indices.get('Giriş Tarihi', -1)] if 'Giriş Tarihi' in header_indices else None
                
                if not all([tc_kimlik_no, ad, soyad, giris_tarihi_raw]):
                    self.stdout.write(self.style.WARNING(f"Uyarı (Satır {row_idx}): Eksik temel veri (T.C. No, Ad, Soyad veya Giriş Tarihi). Bu satır atlandı."))
                    skipped_count += 1
                    continue

                # TC Kimlik Numarası benzersiz olduğu için mevcut kaydı bulmaya çalış
                misafir, created = Misafir.objects.get_or_create(
                    tc_kimlik_no=tc_kimlik_no,
                    defaults={
                        'ad': ad,
                        'soyad': soyad,
                        # Giriş tarihi için varsayılanı burada belirlemeyin, Excel'den gelenle güncelleyin
                        # 'giris_tarihi': timezone.now(), # Eğer Excel'de yoksa veya boşsa
                    }
                )

                if created:
                    imported_count += 1
                else:
                    updated_count += 1
                    self.stdout.write(self.style.WARNING(f"Uyarı (Satır {row_idx}): T.C. Kimlik No '{tc_kimlik_no}' ile mevcut kayıt bulundu. Güncelleniyor..."))

                # Alanları güncelle/ata
                for excel_col_name, model_field_name in column_mapping.items():
                    if excel_col_name in header_indices:
                        value = row_data[header_indices[excel_col_name]]
                        
                        # Tarih/Saat alanları için özel işlem (DateTimeField)
                        if model_field_name in ['giris_tarihi', 'cikis_tarihi']:
                            if value:
                                naive_dt = None
                                if isinstance(value, date) and not isinstance(value, datetime):
                                    # Excel'den gelen sadece tarih ise (datetime.date), bunu datetime'a çevir
                                    naive_dt = datetime(value.year, value.month, value.day)
                                elif isinstance(value, datetime):
                                    naive_dt = value # Zaten datetime objesi
                                else:
                                    # Eğer string olarak gelirse (Excel formatına göre parse etmeniz gerekebilir)
                                    # Örneğin: naive_dt = datetime.strptime(str(value), '%d.%m.%Y %H:%M:%S')
                                    self.stdout.write(self.style.WARNING(f"Uyarı (Satır {row_idx}): '{excel_col_name}' sütunundaki tarih/saat formatı beklenmedik: '{value}'. Boş bırakıldı."))
                                    setattr(misafir, model_field_name, None)
                                    continue
                                
                                if naive_dt and timezone.is_naive(naive_dt):
                                    setattr(misafir, model_field_name, timezone.make_aware(naive_dt, timezone.get_current_timezone()))
                                else:
                                    setattr(misafir, model_field_name, naive_dt) # Zaten aware ise veya None ise
                            else:
                                setattr(misafir, model_field_name, None) # Boş ise None yap
                        
                        # Doğum tarihi için (models.DateField olduğu için timezone gerekmez)
                        elif model_field_name == 'dogum_tarihi':
                            if value:
                                if isinstance(value, date):
                                    setattr(misafir, model_field_name, value)
                                else:
                                    # String gelirse parse etmeyi deneyin (örn: '01.01.2000' formatı için)
                                    # try:
                                    #    setattr(misafir, model_field_name, datetime.strptime(str(value), '%d.%m.%Y').date())
                                    # except ValueError:
                                    self.stdout.write(self.style.WARNING(f"Uyarı (Satır {row_idx}): 'Doğum Tarihi' formatı beklenmedik: '{value}'. Boş bırakıldı."))
                                    setattr(misafir, model_field_name, None)
                            else:
                                setattr(misafir, model_field_name, None)

                        # Durum bilgisi için (choices alanları büyük/küçük harf duyarlı olabilir)
                        elif model_field_name == 'durum':
                            if value:
                                value_upper = str(value).strip().upper()
                                if value_upper in [choice[0] for choice in Misafir.DURUM_SECENEKLERI]:
                                    setattr(misafir, model_field_name, value_upper)
                                else:
                                    self.stdout.write(self.style.WARNING(f"Uyarı (Satır {row_idx}): Geçersiz 'Durum Bilgisi': '{value}'. 'AKTIF' olarak ayarlandı."))
                                    setattr(misafir, model_field_name, 'AKTIF') # Varsayılan atama
                            else:
                                setattr(misafir, model_field_name, 'AKTIF') # Boş ise varsayılan

                        # Yatak Numarası için (Foreign Key)
                        elif model_field_name == 'yatak_no':
                            if value:
                                try:
                                    # Yatağı ID'si veya numarası ile bulabilirsiniz
                                    yatak_obj = Yatak.objects.get(numara=str(value).strip()) # Yatak modelinizde 'numara' alanı olduğunu varsaydık
                                    setattr(misafir, 'yatak_no', yatak_obj)
                                except Yatak.DoesNotExist:
                                    self.stdout.write(self.style.WARNING(f"Uyarı (Satır {row_idx}): Yatak Numarası '{value}' bulunamadı. Bu misafire yatak atanmadı."))
                                    setattr(misafir, 'yatak_no', None)
                            else:
                                setattr(misafir, 'yatak_no', None)

                        # Sosyal Güvence için (Foreign Key)
                        elif model_field_name == 'sosyal_guvence':
                            if value:
                                try:
                                    # Sosyal Güvenceyi adına göre bulabilirsiniz
                                    sosyal_guvence_obj = SosyalGuvence.objects.get(name=str(value).strip()) # SosyalGuvence modelinizde 'name' alanı olduğunu varsaydık
                                    setattr(misafir, 'sosyal_guvence', sosyal_guvence_obj)
                                except SosyalGuvence.DoesNotExist:
                                    self.stdout.write(self.style.WARNING(f"Uyarı (Satır {row_idx}): Sosyal Güvence '{value}' bulunamadı. Bu misafire sosyal güvence atanmadı."))
                                    setattr(misafir, 'sosyal_guvence', None)
                            else:
                                setattr(misafir, 'sosyal_guvence', None)

                        # Diğer basit metin alanları (ForeignKey ve ImageField hariç)
                        elif model_field_name not in ['fotograf']:
                            setattr(misafir, model_field_name, str(value).strip() if value is not None else '')

                # Misafir objesini kaydet (get_or_create ile zaten oluşturulduysa, yapılan değişiklikleri güncellemek için save çağırılır)
                misafir.save()
                    
            except Exception as e:
                self.stdout.write(self.style.ERROR(f"Hata (Satır {row_idx}): '{row_data}' verisi işlenirken bir hata oluştu: {e}"))
                skipped_count += 1
                # Hata izini görmek için uncomment edin:
                # import traceback
                # traceback.print_exc(file=sys.stdout)

        self.stdout.write(self.style.SUCCESS(f"\nVeri aktarımı tamamlandı!"))
        self.stdout.write(self.style.SUCCESS(f"Toplam aktarılan yeni kayıt: {imported_count}"))
        self.stdout.write(self.style.SUCCESS(f"Toplam güncellenen kayıt: {updated_count}"))
        self.stdout.write(self.style.WARNING(f"Toplam atlanan kayıt: {skipped_count}"))